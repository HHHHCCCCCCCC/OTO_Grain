from ctypes import *
import openpyxl
import datetime
from pre_spectrum_Grain import pretreatment as pre
import numpy as np
import math

class IRConfig():
    def __init__(self):
        self.ComFormat = 'M{:0>6s}{:1s}{:0>2s}{:0>2s}'
        self.ComParity = {'None': '0', 'Odd': '1', 'Even': '2'}
        self.wb = openpyxl.load_workbook('config_Grain.xlsx')
        self.Work_Excel = self.wb.active
        self.Work_Excel = self.wb['光谱仪设置']
        self.ModbusConfig = self.ComFormat.format(str(self.Work_Excel.cell(3, 2).value),
                                    self.ComParity[str(self.Work_Excel.cell(5, 2).value)],
                                    str(self.Work_Excel.cell(8, 2).value), str(self.Work_Excel.cell(9, 2).value))
        self.ScanTimes = (self.Work_Excel.cell(13, 2).value)                        # 扫描次数
        self.IntergrationTime_back = int(self.Work_Excel.cell(14, 2).value)              # 积分时间 （us）
        self.IntergrationTime_Grain = int(self.Work_Excel.cell(14, 3).value)              # 积分时间 （us）
        print(self.IntergrationTime_back)
        self.LoopTime = int(self.Work_Excel.cell(15, 2).value)                           # 间隔时间
        self.TemperatureSet = self.Work_Excel.cell(16, 2).value                     # 温度设置
        if str(self.Work_Excel.cell(18, 2).value) == "ON":                          # 制冷启动
            self.TECONOFF = 1
        else:
            self.TECONOFF = 0
        self.Pre_Function = [''] * 10
        self.beta1 = [0] * 1844
        self.beta2 = [0] * 1844
        self.beta3 = [0] * 1844
        self.beta4 = [0] * 1844
        self.beta5 = [0] * 1844
        self.MA_WinSize = 0
        self.SG_WinLen = 0
        self.SG_PolyOrder = 0
        self.spline_xmin = 0
        self.spline_numbers = 0
        self.spline_xmax = 0
        Pre_Cnt = 0
        while True:
            if str(self.Work_Excel.cell(21 + Pre_Cnt, 1).value) != 'end':
                self.Pre_Function[Pre_Cnt] = str(self.Work_Excel.cell(21 + Pre_Cnt, 1).value)
                if self.Pre_Function[Pre_Cnt] == 'smoothMA':
                    self.MA_WinSize = int(self.Work_Excel.cell(21 + Pre_Cnt, 2).value)
                elif self.Pre_Function[Pre_Cnt] == 'Diff':
                    self.SG_WinLen = int(self.Work_Excel.cell(21 + Pre_Cnt, 2).value)
                    self.SG_PolyOrder = int(self.Work_Excel.cell(21 + Pre_Cnt, 3).value)
                elif self.Pre_Function[Pre_Cnt] == 'Spline':
                    self.spline_xmin = int(self.Work_Excel.cell(21 + Pre_Cnt, 2).value)
                    self.spline_xmax = int(self.Work_Excel.cell(21 + Pre_Cnt, 3).value)
                    self.spline_numbers = int(self.Work_Excel.cell(21 + Pre_Cnt, 4).value)
                Pre_Cnt = Pre_Cnt + 1
            else:
                break

        self.Head_Information = [
            ['#######%Modbus协议设置###############'],
            ['1.串口号：', str(self.Work_Excel.cell(2, 2).value)],
            ['2.波特率：', str(self.Work_Excel.cell(3, 2).value)],
            ['3.数据位：',  str(self.Work_Excel.cell(4, 2).value)],
            ['4.校验位：',  str(self.Work_Excel.cell(5, 2).value)],
            ['5.停止位：',  str(self.Work_Excel.cell(6, 2).value)],
            ['6.超时时间',  str(self.Work_Excel.cell(7, 2).value)],
            ['7.设备号：',  str(self.Work_Excel.cell(8, 2).value)],
            ['8.寄存器开始：', str(self.Work_Excel.cell(9, 2).value)],
            ['9.寄存器个数：', str(self.Work_Excel.cell(10, 2).value)],
            [' '],
            ['#########%光谱仪设置：第一个是扫描次数，每次都是6次平均后的结果并保留#########'],
            ['1.扫描次数：', str(self.Work_Excel.cell(13, 2).value)],
            ['2. 积分时间（us)', str(self.Work_Excel.cell(14, 2).value)],
            ['3.间隔时间（ms)：', str(self.Work_Excel.cell(15, 2).value)],
            ['4.探测器温度设置（℃)：:', str(self.Work_Excel.cell(16, 2).value)],
            ['5.光源工作模式：（LC/LO)', str(self.Work_Excel.cell(17, 2).value)],
            [' '],
            ['#########%Pretreat      """光谱预处理按下列顺序执行操作:标准化；归一化；求导；差分···以end为标记结束"""#########']
        ]
        self.sheet2 = self.wb['加权系数']
        for i in range(0, 1844):  #1844个波长,另加一个beta常量
            self.beta1[i] = float(self.sheet2.cell(1 + i, 2).value)
            self.beta2[i] = float(self.sheet2.cell(1 + i, 3).value)
            self.beta3[i] = float(self.sheet2.cell(1 + i, 4).value)
            self.beta4[i] = float(self.sheet2.cell(1 + i, 5).value)
            self.beta5[i] = float(self.sheet2.cell(1 + i, 6).value)
        self.wb.close()

class Device:
    def __init__(self):
        self.VID = 1592
        self.PID = 2732
        self.errStatus = 0
        self.background_intensity = []*249
        self.pre = pre()
        self.OTOdll = CDLL("UserApplication.dll")
        self.LoadConfig = IRConfig()
        self.intFramesize = self.GetFramesize()
        # self.IntergrationTime_Grain = 1000
        # self.current = str(datetime.datetime.now().replace(microsecond=0))
        # self.current = self.current.replace(':', '-')
        # self.Dayintime = self.current


    def IsConnected(self):
        #Check how many device is connected with PC.
        self.intDeviceamout = c_int(0)
        self.OTOdll.UAI_SpectrometerGetDeviceAmount(self.VID,self.PID,byref(self.intDeviceamout))
        return self.intDeviceamout.value

    def Open(self):
        #Open Device
        self.DeviceHandle = c_int(0)
        self.OTOdll.UAI_SpectrometerOpen(0,byref(self.DeviceHandle),self.VID,self.PID)
        return self.DeviceHandle.value

    def GetFramesize(self):
        self.Open()
        #Get Framesize
        self.intFramesize = c_int(0)
        self.OTOdll.UAI_SpectromoduleGetFrameSize(self.DeviceHandle,byref(self.intFramesize))
        print("Device framesize:")
        return self.intFramesize

    def GetModuleName(self):
        #Get Module name
        self.charModulename = create_string_buffer(16)
        self.OTOdll.UAI_SpectrometerGetModelName(self.DeviceHandle,byref(self.charModulename))
        print("Module name:")
        return repr(self.charModulename.value)

    def GetSerilaNum(self):
        #Get Serial number
        self.charSerialnumber = create_string_buffer(16)
        self.OTOdll.UAI_SpectrometerGetSerialNumber(self.DeviceHandle,byref(self.charSerialnumber))
        print("Serial number:")
        return repr(self.charSerialnumber.value)

    def GetLambda(self):
        self.TempLambda = (c_float*self.intFramesize.value)()
        #Get wavelength
        self.OTOdll.UAI_SpectrometerWavelengthAcquire(self.DeviceHandle,byref(self.TempLambda))
        self.Lambda = []
        for i in range(0, self.intFramesize.value):
            self.Lambda.append(self.TempLambda[i])
        return self.Lambda

    def TEC(self):
        # self.GetTECOnOff = c_int(0)
        self.OTOdll.UAI_SpectrometerSetTECOnOff(self.DeviceHandle, self.LoadConfig.TECONOFF)
        # self.OTOdll.UAI_SpectrometerGetTECOnOff(self.DeviceHandle, byref(self.GetTECOnOff))
        self.OTOdll.UAI_SpectrometerSetTECTargetTemperature(self.DeviceHandle, c_float(self.LoadConfig.TemperatureSet))
        # self.OTOdll.UAI_SpectrometerSetTECDAC(self.DeviceHandle, 0xff)

    def GetIntensity_Grain(self,IntergrationTime_Grain):
        self.TempIntensity = (c_float * self.intFramesize.value)()
        self.OTOdll.UAI_SpectrometerDataOneshots(self.DeviceHandle,IntergrationTime_Grain,byref(self.TempIntensity),int(self.LoadConfig.ScanTimes))
        # Do Background
        self.OTOdll.UAI_BackgroundRemoveWithAVG(self.DeviceHandle, self.intFramesize.value, byref(self.TempIntensity))
        # Do Linearity
        self.OTOdll.UAI_LinearityCorrection(self.DeviceHandle, self.intFramesize.value, byref(self.TempIntensity))
        # Get Intensity
        self.Intensity = []
        for i in range(0, self.intFramesize.value):
            self.Intensity.append((self.TempIntensity[i])/IntergrationTime_Grain)
        return self.Intensity

    def GetIntensity_back(self):
        self.TempIntensity = (c_float * self.intFramesize.value)()
        #Get Intensity
        self.OTOdll.UAI_SpectrometerDataOneshots(self.DeviceHandle,self.LoadConfig.IntergrationTime_back,byref(self.TempIntensity),int(self.LoadConfig.ScanTimes))
        # Do Background
        self.OTOdll.UAI_BackgroundRemoveWithAVG(self.DeviceHandle, self.LoadConfig.IntergrationTime_back, byref(self.TempIntensity))
        # Do Linearity
        self.OTOdll.UAI_LinearityCorrection(self.DeviceHandle, self.LoadConfig.IntergrationTime_back, byref(self.TempIntensity))
        self.Intensity = []
        for i in range(0, self.intFramesize.value):
            self.Intensity.append((self.TempIntensity[i])/self.LoadConfig.IntergrationTime_back)
        return self.Intensity



    ###############以上为预处理方法#################
    def SaveExcel(self, excel_name, indesity_data):
        wb = openpyxl.Workbook()
        Work_Excel = wb.active
        for j in range(0, len(self.LoadConfig.Head_Information)):
            Work_Excel.append(self.LoadConfig.Head_Information[j])  # 把每一行append到worksheet中
        for j in range(0, len(self.LoadConfig.Pre_Function)):
            Work_Excel.append([self.LoadConfig.Pre_Function[j]])  # 把每一行append到worksheet中
        for j in range(0, len(self.WaveLenth)):
            Work_Excel.append([self.WaveLenth[j]] + [indesity_data[j]])
        wb.save(excel_name)

    def SaveBeijingExcel(self, excel_name, indesity_data):
        wb = openpyxl.Workbook()
        Work_Excel = wb.active
        self.WaveLenth = self.GetLambda()
        for j in range(0, len(self.WaveLenth)):
            Work_Excel.append([self.WaveLenth[j]] + [indesity_data[j]])
        wb.save(excel_name)

    def WriteDataToExcl(self, data,background_intensity,current):
        # self.current = str(datetime.datetime.now().replace(microsecond=0))
        # self.current = self.current.replace(':', '-')
        excel_filename2 = "D:/Pukon/Pukon_" + str('谷物透射光谱') + "/" + current + '透射光谱' + ".xlsx"
        excel_filename3 = "D:/Pukon/Pukon_" + str('谷物原始吸收光谱') + "/" + current + '原始吸收光谱' + ".xlsx"
        excel_filename4 = "D:/Pukon/Pukon_" + str('谷物预处理吸收光谱') + "/" + current + '预处理吸收光谱' + ".xlsx"
        self.WaveLenth = self.GetLambda()
        self.SaveExcel(excel_filename2, data)
        self.Absorb = [0]*1595
        self.background_intensity = background_intensity
        for i in range(0, len(data)):
            self.Absorb[i] = math.log10(self.background_intensity[i]/data[i])
        self.SaveExcel(excel_filename3, self.Absorb)
        # for i in range(0, len(self.LoadConfig.Pre_Function)):
        #     if self.LoadConfig.Pre_Function[i] != '':
        #         excel_filename = "D:/Pukon/Pukon_" + str(self.LoadConfig.Pre_Function[i]) + "/" + self.Dayintime + '-' \
        #                          + str(self.LoadConfig.Pre_Function[i]) + ".xlsx"
        #         if str(self.LoadConfig.Pre_Function[i]) == 'SmoothSG':
        #             self.Intensities_Smooth_Temp = self.smoothSG(data.copy(), self.LoadConfig.SG_WinLen, self.LoadConfig.SG_PolyOrder)
        #             self.SaveExcel(excel_filename, self.Intensities_Smooth_Temp)
        #         elif str(self.LoadConfig.Pre_Function[i]) == 'SmoothMA':
        #             self.Intensities_Smooth_Temp = self.smoothMA(data.copy(), self.LoadConfig.MA_WinSize)
        #             self.Intensities_Smooth_Temp = self.Intensities_Temp.tolist()
        #             self.SaveExcel(excel_filename, self.Intensities_Smooth_Temp)
        #         #
        #         if str(self.LoadConfig.Pre_Function[i]) == 'MaxNorm':
        #             self.Intensities_Norm_Temp = self.MaxNorm(data.copy())
        #             self.Intensities_Norm_Temp = self.Intensities_Norm_Temp.tolist()
        #             self.SaveExcel(excel_filename, self.Intensities_Norm_Temp)
        #         #
        #         if str(self.LoadConfig.Pre_Function[i]) == 'Diff1':
        #             self.Intensities_Diff_Temp = self.D1(self.WaveLenth.copy(), data.copy())
        #             self.SaveExcel(excel_filename, self.Intensities_Diff_Temp)
        #         # elif str(self.LoadConfig.Pre_Function[i]) == 'Diff2':
        #         #     self.Intensities_Diff_Temp = self.D2(self.WaveLenth.copy(), self.Intensities_Norm_Temp.copy())
        #         #     self.SaveExcel(excel_filename, self.Intensities_Diff_Temp)
        #         #
        #         if str(self.LoadConfig.Pre_Function[i]) == 'SNV':
        #             self.Intensities_SNV_Temp = self.SNV(data.copy())  # 注意，此处必须加入.copy(),否者会改变传递到函数中的列表
        #             self.SaveExcel(excel_filename, self.Intensities_SNV_Temp)
        #         #
        #         # if str(self.LoadConfig.Pre_Function[i]) == 'MSC':
        #         #     self.Intensities_Correct_Temp = self.MSC(self.Intensities_SNV_Temp.copy())
        #         #     self.SaveExcel(excel_filename, self.Intensities_Correct_Temp)
        #         # elif str(self.LoadConfig.Pre_Function[i]) == 'MC':
        #         #     self.Intensities_Correct_Temp = self.MeanCenter(self.Intensities_SNV_Temp.copy())
        #         #     self.SaveExcel(excel_filename, self.Intensities_Correct_Temp)
        #     # else:
        for i in range(0, len(self.LoadConfig.Pre_Function)):
            if self.LoadConfig.Pre_Function[i] != '':
                print(self.LoadConfig.Pre_Function[i])
                if i==0:
                    self.Intensities_Temp = self.Absorb
                else:
                    self.Intensities_Temp = self.Intensities_Temp
                if str(self.LoadConfig.Pre_Function[i]) == 'SNV':
                    self.Intensities_Temp = self.pre.SNV(self.Intensities_Temp)
                elif str(self.LoadConfig.Pre_Function[i]) == 'Diff':
                    self.Intensities_Temp = self.pre.Diff(self.Intensities_Temp, self.LoadConfig.SG_WinLen,self.LoadConfig.SG_PolyOrder)

                elif str(self.LoadConfig.Pre_Function[i]) == 'smoothMA':
                    self.Intensities_Temp = self.pre.smoothMA(self.Intensities_Temp, self.LoadConfig.MA_WinSize)

                elif str(self.LoadConfig.Pre_Function[i]) == 'MaxNorm':
                    self.Intensities_Temp = self.pre.MaxNorm(self.Intensities_Temp)

                elif str(self.LoadConfig.Pre_Function[i]) == 'MaxMinNormalize':
                    self.Intensities_Temp = self.pre.MaxMinNormalize(self.Intensities_Temp)

                elif str(self.LoadConfig.Pre_Function[i]) == 'MeanCenter':
                    self.Intensities_Temp = self.pre.MeanCenter(self.Intensities_Temp)

                elif str(self.LoadConfig.Pre_Function[i]) == 'vector_Normalize':
                    self.Intensities_Temp = self.pre.vector_Normalize(self.Intensities_Temp)

                # if str(self.LoadConfig.Pre_Function[i]) == 'Spline':
                #     print(123)
                #     self.WaveLenth ,self.Intensities = self.pre.Spline( self.WaveLenth,self.Intensities ,self.LoadConfig.spline_xmin ,self.LoadConfig.spline_xmax , self.LoadConfig.spline_numbers)
        self.Intensities_Temp = np.array(self.Intensities_Temp).flatten()
        self.Intensities_Temp = self.Intensities_Temp.tolist()
        self.SaveExcel(excel_filename4, self.Intensities_Temp)

    def SaveJiancezhi(self, data):
        wb =openpyxl.load_workbook("D:/Pukon/" + "Zhijian/" +  "质检" + ".xlsx")
        Work_Excel = wb.active
        Work_Excel.append(data)
        excel_filename = "D:/Pukon/" + "Zhijian/" +  "质检" + ".xlsx"
        wb.save(excel_filename)